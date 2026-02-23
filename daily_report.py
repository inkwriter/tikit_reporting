#!/usr/bin/env python3
"""
IT Daily Report Generator
Generates: IT_Daily_Report.xlsx with 4 tabs
"""

import os
import pandas as pd

# All store names for tracking
ALL_STORES = [
    'Jenkins', 'Neon', 'Harlan 1', 'Harlan 2', 'Hyden', 'PMM', 'Isom', 
    'Whitesburg', 'Hazard 2', 'PMG', 'Ermine (Arbys)', 'Hindman 2 (Arbys)', 
    'Hindman 1', 'Martin', 'Jackson', 'Hazard 3', 'Dryfork', 'Pound', 
    'Nicholasville', 'Catnip', 'Marrowbone', 'Elkhorn City', 'Chloe', 
    'Caney', 'Belfrey', 'Phelps', 'Virgie', 'Harold', 'Allen', 'Goody', 
    'Zebulon', 'Pikeville South', 'North', 'Prestonsburg 1', 'Ivel', 
    'Justiceville', 'Salyersville', 'Grundy', 'West Liberty', 
    'Prestonsburg 2', 'Prestonsburg 23'
]

def load_and_process_data():
    """Load and process ticket data"""
    print("\nLoading data...")
    active = pd.read_csv('active_tickets.csv')
    closed = pd.read_csv('closed_tickets.csv')
    
    # Filter to IT Helpdesk team only
    if 'Team' in active.columns:
        active = active[active['Team'] == 'IT Helpdesk']
    if 'Team' in closed.columns:
        closed = closed[closed['Team'] == 'IT Helpdesk']
    
    print(f"✓ Filtered to IT Helpdesk team only")
    
    # Filter closed to last 7 calendar days (midnight 7 days ago)
    closed['Last Modified Date'] = pd.to_datetime(closed['Last Modified Date'], errors='coerce', utc=True)
    seven_days_ago = (pd.Timestamp.now(tz='UTC') - pd.Timedelta(days=7)).normalize()
    closed = closed[closed['Last Modified Date'] >= seven_days_ago]
    
    # Remove timezone info from all datetime columns to avoid Excel issues
    for col in closed.columns:
        if pd.api.types.is_datetime64_any_dtype(closed[col]):
            closed[col] = closed[col].dt.tz_localize(None)
    
    for col in active.columns:
        if pd.api.types.is_datetime64_any_dtype(active[col]):
            active[col] = active[col].dt.tz_localize(None)
    
    # Treat "Resolved" status as closed too - pull from active
    active_copy = active.copy()
    resolved = active_copy[active_copy['Status'].str.lower().str.contains('resolved', na=False)]
    active = active_copy[~active_copy['Status'].str.lower().str.contains('resolved', na=False)]
    
    # Combine resolved with closed
    if len(resolved) > 0:
        closed = pd.concat([closed, resolved], ignore_index=True)
    
    # Add Status_Type BEFORE combining
    active['Status_Type'] = 'Active'
    closed['Status_Type'] = 'Closed'
    
    print(f"✓ Active tickets: {len(active)}")
    print(f"✓ Closed/Resolved tickets (last 7 days): {len(closed)}")
    
    return active, closed

def generate_it_daily_report(active, closed):
    """Generate IT Daily Report with 4 tabs"""
    print("\n" + "=" * 50)
    print("IT DAILY REPORT")
    print("=" * 50)
    
    filename = 'IT_Daily_Report.xlsx'

    from openpyxl.styles import PatternFill, Font, Alignment
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=False)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # ========== TAB 1: READ EXISTING HISTORY FIRST (before opening writer) ==========
    print("\n📊 Creating Tab 1: Assignee Summary...")

    today_str = pd.Timestamp.now().strftime('%m-%d-%Y')

    all_for_summary = pd.concat([active, closed], ignore_index=True)
    assigned = all_for_summary[all_for_summary['Assignee'].notna() & (all_for_summary['Assignee'] != '')]

    today_rows = []
    for assignee in sorted(assigned['Assignee'].unique()):
        assignee_tickets = assigned[assigned['Assignee'] == assignee]
        active_count = len(assignee_tickets[assignee_tickets['Status_Type'] == 'Active'])
        closed_count = len(assignee_tickets[assignee_tickets['Status_Type'] == 'Closed'])
        today_rows.append({
            'Date': today_str,
            'Assignee': assignee,
            'Active Tickets': active_count,
            'Closed (Last 7 Days)': closed_count,
            'Total': active_count + closed_count
        })
    today_df = pd.DataFrame(today_rows)

    # Read existing history BEFORE opening the writer (opening writer wipes the file)
    existing_history = None
    if os.path.exists(filename):
        try:
            existing_history = pd.read_excel(filename, sheet_name='Assignee Summary')
            if 'Date' not in existing_history.columns:
                existing_history.insert(0, 'Date', '')
            print(f"✔ Loaded existing history ({len(existing_history)} rows)")
        except Exception as e:
            print(f"  Could not load existing history: {e}")
            existing_history = None

    # Now build the full summary dataframe
    if existing_history is not None and len(existing_history) > 0:
        blank_row = pd.DataFrame([{col: '' for col in existing_history.columns}])
        summary_df = pd.concat([existing_history, blank_row, today_df], ignore_index=True)
    else:
        summary_df = today_df

    # NOW open the writer (this is when the file gets wiped/created)
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    summary_df.to_excel(writer, sheet_name='Assignee Summary', index=False)

    # ========== TAB 2: ACTIVE TICKETS (SEPARATED BY ASSIGNEE) ==========
    print("📋 Creating Tab 2: Active Tickets...")
    
    active_sheet_data = []
    
    # Add unassigned section first
    unassigned = active[active['Assignee'].isna() | (active['Assignee'] == '')]
    if len(unassigned) > 0:
        active_sheet_data.append(['UNASSIGNED', '', '', '', '', '', '', ''])
        active_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Collaborators', 'Created Date', 'Assignee'])
        for _, ticket in unassigned.iterrows():
            created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
            active_sheet_data.append([
                ticket.get('Id', ''),
                ticket.get('Request', ''),
                ticket.get('Requester', ''),
                ticket.get('Priority', ''),
                ticket.get('Category', ''),
                ticket.get('Collaborators', ''),
                created_date,
                'UNASSIGNED'
            ])
        active_sheet_data.append(['', '', '', '', '', '', '', ''])
    
    # Add each assignee's section
    assigned_active = active[active['Assignee'].notna() & (active['Assignee'] != '')]
    for assignee in sorted(assigned_active['Assignee'].unique()):
        assignee_tickets = assigned_active[assigned_active['Assignee'] == assignee]
        active_sheet_data.append([assignee, '', '', '', '', '', '', ''])
        active_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Collaborators', 'Created Date', 'Assignee'])
        for _, ticket in assignee_tickets.iterrows():
            created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
            active_sheet_data.append([
                ticket.get('Id', ''),
                ticket.get('Request', ''),
                ticket.get('Requester', ''),
                ticket.get('Priority', ''),
                ticket.get('Category', ''),
                ticket.get('Collaborators', ''),
                created_date,
                assignee
            ])
        active_sheet_data.append(['', '', '', '', '', '', '', ''])
    
    active_df = pd.DataFrame(active_sheet_data)
    active_df.to_excel(writer, sheet_name='Active Tickets', index=False, header=False)
    
    worksheet = writer.sheets['Active Tickets']
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 60
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 18
    
    for row in range(1, len(active_sheet_data) + 1):
        for col in range(1, 9):
            cell = worksheet.cell(row, col)
            cell.alignment = wrap_align if col == 2 else top_left_align
        
        cell_value = worksheet.cell(row, 8).value
        if cell_value == 'UNASSIGNED' or (row > 1 and worksheet.cell(row, 1).value == 'UNASSIGNED'):
            for col in range(1, 9):
                cell = worksheet.cell(row, col)
                cell.fill = yellow_fill
                cell.alignment = wrap_align if col == 2 else top_left_align
        
        if row > 1 and row < len(active_sheet_data) and worksheet.cell(row + 1, 1).value == 'Ticket ID':
            worksheet.cell(row, 1).font = bold_font
    
    # ========== TAB 3: CLOSED TICKETS (SEPARATED BY ASSIGNEE) ==========
    print("✅ Creating Tab 3: Closed Tickets...")
    
    closed_sheet_data = []
    assigned_closed = closed[closed['Assignee'].notna() & (closed['Assignee'] != '')]
    
    if len(assigned_closed) == 0:
        closed_sheet_data.append(['No closed tickets in the last 7 days'])
    else:
        for assignee in sorted(assigned_closed['Assignee'].unique()):
            assignee_tickets = assigned_closed[assigned_closed['Assignee'] == assignee]
            closed_sheet_data.append([assignee, '', '', '', '', '', '', '', ''])
            closed_sheet_data.append(['Ticket ID', 'Request', 'Requester', 'Priority', 'Category', 'Collaborators', 'Created Date', 'Last Modified Date', 'Assignee'])
            for _, ticket in assignee_tickets.iterrows():
                created_date = str(ticket.get('Created Date', ''))[:19] if pd.notna(ticket.get('Created Date', '')) else ''
                modified_date = str(ticket.get('Last Modified Date', ''))[:19] if pd.notna(ticket.get('Last Modified Date', '')) else ''
                closed_sheet_data.append([
                    ticket.get('Id', ''),
                    ticket.get('Request', ''),
                    ticket.get('Requester', ''),
                    ticket.get('Priority', ''),
                    ticket.get('Category', ''),
                    ticket.get('Collaborators', ''),
                    created_date,
                    modified_date,
                    assignee
                ])
            closed_sheet_data.append(['', '', '', '', '', '', '', '', ''])
    
    closed_df = pd.DataFrame(closed_sheet_data)
    closed_df.to_excel(writer, sheet_name='Closed Tickets', index=False, header=False)
    
    worksheet_closed = writer.sheets['Closed Tickets']
    worksheet_closed.column_dimensions['A'].width = 10
    worksheet_closed.column_dimensions['B'].width = 60
    worksheet_closed.column_dimensions['C'].width = 20
    worksheet_closed.column_dimensions['D'].width = 12
    worksheet_closed.column_dimensions['E'].width = 20
    worksheet_closed.column_dimensions['F'].width = 20
    worksheet_closed.column_dimensions['G'].width = 20
    worksheet_closed.column_dimensions['H'].width = 20
    worksheet_closed.column_dimensions['I'].width = 18
    
    for row in range(1, len(closed_sheet_data) + 1):
        for col in range(1, 10):
            cell = worksheet_closed.cell(row, col)
            cell.alignment = wrap_align if col == 2 else top_left_align
        if row > 1 and row < len(closed_sheet_data) and worksheet_closed.cell(row + 1, 1).value == 'Ticket ID':
            worksheet_closed.cell(row, 1).font = bold_font
    
    # ========== TAB 4: STORE COUNTS ==========
    print("🏪 Creating Tab 4: Store Counts...")
    
    all_for_stores = pd.concat([active, closed], ignore_index=True)
    store_count_data = []
    for store in ALL_STORES:
        count = len(all_for_stores[all_for_stores['Requester'].str.contains(store, case=False, na=False)])
        store_count_data.append({'Store Name': store, 'Number of Tickets': count})
    
    store_df = pd.DataFrame(store_count_data)
    store_df = store_df.sort_values('Number of Tickets', ascending=False)
    store_df.to_excel(writer, sheet_name='Store Counts', index=False)
    
    worksheet_stores = writer.sheets['Store Counts']
    worksheet_stores.column_dimensions['A'].width = 25
    worksheet_stores.column_dimensions['B'].width = 18
    
    for row in range(1, len(store_df) + 2):
        for col in range(1, 3):
            worksheet_stores.cell(row, col).alignment = top_left_align
    for col in range(1, 3):
        worksheet_stores.cell(1, col).font = bold_font
    
    writer.close()
    print(f"✓ Saved: {filename}")

def main():
    print("=" * 50)
    print("IT DAILY REPORT GENERATOR")
    print("=" * 50)
    
    active, closed = load_and_process_data()
    generate_it_daily_report(active, closed)
    
    print("\n" + "=" * 50)
    print("✅ REPORT GENERATED!")
    print("=" * 50)
    print("\nFile created: IT_Daily_Report.xlsx\n")

if __name__ == "__main__":
    main()
